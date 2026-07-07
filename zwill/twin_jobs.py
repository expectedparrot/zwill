from __future__ import annotations

import json
import random
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

from .errors import ZwillError
from .jsonlio import read_jsonl
from .twin import digital_twin_job_id_from_job, select_context_questions


@dataclass(frozen=True)
class DigitalTwinJobBuilderDeps:
    require_survey: Callable[[str], Path]
    selected_question_names: Callable[[Any, list[dict[str, Any]]], list[str]]
    respondent_selection: Callable[[Any, list[str]], list[str]]
    context_question_options: Callable[[dict[str, Any]], list[str]]
    context_path: Callable[[Path], Path]
    load_edsl_job_classes: Callable[[], tuple[Any, Any, Any, Any, Any, Any, Any]]
    load_twin_material: Callable[[Any], list[dict[str, Any]]]
    selected_agent_material_kinds: Callable[[Any], set[str]]
    selected_agent_material_tags: Callable[[Any], set[str]]
    select_agent_material: Callable[[Path, list[str], Any], list[dict[str, Any]]]
    format_agent_material: Callable[[list[dict[str, Any]], int | None], str]
    matching_twin_material: Callable[..., list[dict[str, Any]]]
    format_twin_material: Callable[[list[dict[str, Any]], int | None], str]
    twin_material_paths: Callable[[Any], list[str]]
    option_key: Callable[[int], str]
    parse_model_params: Callable[[Any], dict[tuple[str | None, str | None], dict[str, Any]]]
    parse_model_specs: Callable[[Any], list[tuple[str, str | None]]]
    model_kwargs_for: Callable[[str, str | None, dict[tuple[str | None, str | None], dict[str, Any]]], dict[str, Any]]


def slug_id(value: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value)).strip("_")
    return slug or "item"


def read_json(path: Path) -> Any:
    return json.loads(path.read_text())


FIELD_PLACEHOLDER_RE = re.compile(r"\[Field-([^\]]+)\]")


def expand_question_text_fields(
    text: str,
    respondent_answers: dict[str, str],
    question_by_name: dict[str, dict[str, Any]],
) -> str:
    label_to_question: dict[str, str] = {}
    for question_name, question in question_by_name.items():
        labels = {
            str(question_name),
            str(question_name).lower(),
            str(question.get("question_text") or "").strip(),
            str(question.get("source", {}).get("raw_label") or "").strip(),
        }
        for label in labels:
            if not label:
                continue
            normalized = re.sub(r"[^a-z0-9]+", "", label.lower())
            if normalized:
                label_to_question.setdefault(normalized, question_name)

    def replacement(match: re.Match[str]) -> str:
        field_name = match.group(1).strip()
        normalized = re.sub(r"[^a-z0-9]+", "", field_name.lower())
        question_name = label_to_question.get(normalized)
        if not question_name:
            return match.group(0)
        answer = respondent_answers.get(question_name)
        return str(answer) if answer is not None else match.group(0)

    return FIELD_PLACEHOLDER_RE.sub(replacement, text)


def selected_heldout_question_names(args: Any, questions: list[dict[str, Any]]) -> list[str]:
    values: list[str] = []
    heldout_question = getattr(args, "heldout_question", None)
    if isinstance(heldout_question, list):
        values.extend(heldout_question)
    elif heldout_question:
        values.append(heldout_question)
    if getattr(args, "heldout_questions", None):
        values.extend(name.strip() for name in args.heldout_questions.split(",") if name.strip())
    if not values:
        raise ZwillError("invalid_input", "--heldout-question is required for twin-probability-job exports.")
    available = [question["question_name"] for question in questions]
    unknown = [name for name in values if name not in available]
    if unknown:
        raise ZwillError(
            "invalid_input",
            "Unknown held-out question selected for digital twin export.",
            context={"unknown_questions": unknown, "available_questions": available},
        )
    deduped = []
    for value in values:
        if value not in deduped:
            deduped.append(value)
    return deduped


def normalize_question_spec(raw: dict[str, Any], *, source_note: str | None = None) -> dict[str, Any]:
    name = str(raw.get("question_name") or raw.get("name") or raw.get("question") or "").strip()
    text = str(raw.get("question_text") or raw.get("text") or "").strip()
    raw_options = raw.get("question_options", raw.get("options", raw.get("option_labels", [])))
    if not isinstance(raw_options, list):
        raise ZwillError("invalid_input", "Question spec options must be a list.", context={"question_name": name})
    options = [str(option).strip() for option in raw_options if str(option).strip()]
    if not name or not text or not options:
        raise ZwillError(
            "invalid_input",
            "Question specs require question_name, question_text, and non-empty question_options.",
            context={"question_name": name, "has_text": bool(text), "option_count": len(options)},
        )
    source = dict(raw.get("source") or {})
    if source_note and not source.get("note"):
        source["note"] = source_note
    return {
        "question_name": name,
        "question_type": raw.get("question_type") or "multiple_choice",
        "question_text": text,
        "question_options": options,
        "source": source,
    }


def read_question_specs(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise ZwillError("not_found", f"Question spec file does not exist: {path}.")
    if path.suffix == ".jsonl":
        raw_rows = read_jsonl(path)
    else:
        raw = read_json(path)
        if isinstance(raw, dict):
            raw_rows = raw.get("questions") or raw.get("question_specs") or []
        elif isinstance(raw, list):
            raw_rows = raw
        else:
            raise ZwillError("invalid_input", "Question spec JSON must be a list or an object with questions.")
    return [normalize_question_spec(row, source_note=f"External question spec: {path}") for row in raw_rows]


def workbook_option_label(raw: Any, answer_value_labels: str) -> str:
    value = str(raw).strip()
    if answer_value_labels == "1=lowest agreement/likelihood/appeal/excitement; 7=highest":
        if value == "1":
            return "1 - lowest agreement/likelihood/appeal/excitement"
        if value == "7":
            return "7 - highest agreement/likelihood/appeal/excitement"
        if value in {"2", "3", "4", "5", "6"}:
            return f"{value} - Likert scale point {value}"
    return value


def read_question_specs_from_workbook(
    path: Path,
    *,
    sheet_name: str,
    question_names: set[str] | None,
    code_column: str,
    text_column: str,
    option_prefix: str,
    labels_column: str,
) -> list[dict[str, Any]]:
    if not path.exists():
        raise ZwillError("not_found", f"Question spec workbook does not exist: {path}.")
    try:
        import openpyxl
    except ImportError as exc:
        raise ZwillError(
            "missing_dependency",
            "openpyxl is required to read question specs from workbooks.",
            hint="Install project dependencies before using --question-specs-workbook.",
        ) from exc
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ZwillError("invalid_input", "Workbook sheet not found.", context={"sheet": sheet_name, "available": wb.sheetnames})
    ws = wb[sheet_name]
    headers = [str(ws.cell(1, column).value or "").strip() for column in range(1, ws.max_column + 1)]
    missing_headers = [name for name in [code_column, text_column] if name not in headers]
    if missing_headers:
        raise ZwillError("invalid_input", "Workbook is missing required question spec columns.", context={"missing": missing_headers})
    specs = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = dict(zip(headers, row))
        name = str(record.get(code_column) or "").strip()
        if not name or (question_names is not None and name not in question_names):
            continue
        answer_value_labels = str(record.get(labels_column) or "").strip() if labels_column in headers else ""
        options = []
        for key in headers:
            if not key.startswith(option_prefix):
                continue
            value = record.get(key)
            if value is None or str(value).strip() == "":
                continue
            options.append(workbook_option_label(value, answer_value_labels))
        specs.append(
            normalize_question_spec(
                {
                    "question_name": name,
                    "question_text": str(record.get(text_column) or "").strip(),
                    "question_options": options,
                    "source": {
                        "raw_id": str(path),
                        "note": f"Workbook {path}, sheet {sheet_name}, row {row_number}.",
                        "answer_value_labels": answer_value_labels or None,
                    },
                }
            )
        )
    if question_names is not None:
        found = {spec["question_name"] for spec in specs}
        missing = sorted(question_names - found)
        if missing:
            raise ZwillError("invalid_input", "Workbook is missing requested question specs.", context={"missing": missing})
    return specs


def extra_heldout_question_specs(args: Any) -> list[dict[str, Any]]:
    specs = []
    if getattr(args, "question_specs", None):
        specs.extend(read_question_specs(Path(args.question_specs)))
    if getattr(args, "question_specs_workbook", None):
        requested = set()
        for value in getattr(args, "heldout_question", None) or []:
            requested.add(str(value))
        if getattr(args, "heldout_questions", None):
            requested.update(item.strip() for item in str(args.heldout_questions).split(",") if item.strip())
        specs.extend(
            read_question_specs_from_workbook(
                Path(args.question_specs_workbook),
                sheet_name=getattr(args, "question_specs_sheet", "Questions"),
                question_names=requested or None,
                code_column=getattr(args, "question_specs_code_column", "Question code"),
                text_column=getattr(args, "question_specs_text_column", "Question text"),
                option_prefix=getattr(args, "question_specs_option_prefix", "Answer option "),
                labels_column=getattr(args, "question_specs_labels_column", "Answer value labels"),
            )
        )
    return specs


def target_specific_leakage_exclusions(args: Any) -> dict[str, set[str]]:
    exclusions: dict[str, set[str]] = defaultdict(set)
    for value in getattr(args, "leakage_exclusion", None) or []:
        text = str(value).strip()
        if not text:
            continue
        if ":" not in text:
            raise ZwillError(
                "invalid_input",
                "Leakage exclusions must use target:question syntax.",
                context={"value": value},
                hint="Example: --leakage-exclusion TATTOO_HAVE_W130:TATTOO_REGRET_W130",
            )
        target, question = [part.strip() for part in text.split(":", 1)]
        if not target or not question:
            raise ZwillError("invalid_input", "Leakage exclusions require both target and question.", context={"value": value})
        exclusions[target].add(question)
    return exclusions


def resolve_leakage_exclusion_patterns(
    raw_exclusions: dict[str, set[str]],
    known_question_names: set[str],
    rank_task_items: dict[str, list[str]],
) -> dict[str, set[str]]:
    """Expand each exclusion pattern into concrete context question names.

    A pattern may be an exact question name, a `rank_task_id` (expands to every
    item in that battery), or a shell-style glob (`prefix*`) matched against the
    known question names. Every pattern must resolve to at least one question.
    """
    import fnmatch

    resolved: dict[str, set[str]] = defaultdict(set)
    unmatched: list[str] = []
    for target, patterns in raw_exclusions.items():
        for pattern in patterns:
            if pattern in rank_task_items:
                resolved[target].update(rank_task_items[pattern])
            elif pattern in known_question_names:
                resolved[target].add(pattern)
            elif any(char in pattern for char in "*?[]"):
                matches = {name for name in known_question_names if fnmatch.fnmatchcase(name, pattern)}
                if matches:
                    resolved[target].update(matches)
                else:
                    unmatched.append(f"{target}:{pattern}")
            else:
                unmatched.append(f"{target}:{pattern}")
    if unmatched:
        raise ZwillError(
            "invalid_input",
            "Leakage exclusions reference unknown context questions, rank tasks, or globs.",
            context={"unmatched": sorted(unmatched)},
            hint="Use target:<question>, target:<rank_task_id>, or a glob like target:q13_message_*.",
        )
    return resolved


def balanced_by_actual(
    respondent_ids: list[str],
    answer_by_respondent: dict[str, dict[str, str]],
    heldout_question: str,
    sample_size: int | None,
    seed: int | None,
) -> list[str]:
    groups: dict[str, list[str]] = defaultdict(list)
    for respondent_id in respondent_ids:
        actual = answer_by_respondent.get(respondent_id, {}).get(heldout_question)
        if actual is not None:
            groups[actual].append(respondent_id)
    if not groups:
        return []
    rng = random.Random(seed)
    for group in groups.values():
        rng.shuffle(group)
    group_count = len(groups)
    if sample_size is None:
        per_group = min(len(group) for group in groups.values())
        remainder = 0
    else:
        per_group = sample_size // group_count
        remainder = sample_size % group_count
    selected = []
    for index, actual in enumerate(sorted(groups)):
        take = per_group + (1 if index < remainder else 0)
        selected.extend(groups[actual][:take])
    return selected


def stratified_by_actual(
    respondent_ids: list[str],
    answer_by_respondent: dict[str, dict[str, str]],
    heldout_question: str,
    sample_size: int | None,
    seed: int | None,
) -> list[str]:
    groups: dict[str, list[str]] = defaultdict(list)
    for respondent_id in respondent_ids:
        actual = answer_by_respondent.get(respondent_id, {}).get(heldout_question)
        if actual is not None:
            groups[actual].append(respondent_id)
    if not groups or sample_size is None:
        return respondent_ids
    rng = random.Random(seed)
    for group in groups.values():
        rng.shuffle(group)
    total = sum(len(group) for group in groups.values())
    allocations = {actual: max(1, round(sample_size * len(group) / total)) for actual, group in groups.items()}
    while sum(allocations.values()) > sample_size:
        largest = max(allocations, key=lambda key: allocations[key])
        allocations[largest] -= 1
    while sum(allocations.values()) < sample_size:
        largest_remainder = max(groups, key=lambda key: len(groups[key]) - allocations.get(key, 0))
        allocations[largest_remainder] += 1
    selected = []
    for actual in sorted(groups):
        selected.extend(groups[actual][: allocations[actual]])
    rng.shuffle(selected)
    return selected


def answer_commonness_by_question(answer_by_respondent: dict[str, dict[str, str]]) -> dict[str, Counter[str]]:
    counts: dict[str, Counter[str]] = defaultdict(Counter)
    for respondent_answers in answer_by_respondent.values():
        for question_name, answer in respondent_answers.items():
            counts[question_name][answer] += 1
    return counts


def answer_commonness_text(question_name: str, answer: str, counts_by_question: dict[str, Counter[str]]) -> str:
    counts = counts_by_question.get(question_name, Counter())
    total = sum(counts.values())
    count = counts.get(answer, 0)
    share = count / total if total else 0.0
    return f"Answer commonness: {count}/{total} respondents ({share:.1%}) gave this answer."


def digital_twin_question_text(prompt_variant: str) -> str:
    base = """You are acting as a digital twin for one survey respondent.

Survey name:
{{ survey_name }}

Survey context:
{{ survey_context }}

Respondent id:
{{ respondent_id }}

Non-survey agent construction material:
{{ agent_material_text }}

Supplemental twin material:
{{ twin_material_text }}

Observed answers from this respondent:
{{ observed_answers_text }}

Held-out question name:
{{ heldout_question_name }}

Held-out question text:
{{ heldout_question_text }}

Held-out response options:
{{ heldout_options_text }}

Use the non-survey agent construction material, supplemental twin material, and observed answers to infer this respondent's probability distribution over the held-out response options.
"""
    if prompt_variant == "answer-commonness-confidence":
        return base + """
When an observed answer includes answer commonness statistics, use those statistics only as context for interpreting that observed answer. They are not statistics for the held-out question unless the held-out question is itself shown as an observed answered question, which standard hold-out exports do not do.

Return only valid JSON. Do not include markdown fences, prose, or comments.

The JSON must have exactly this shape:
{
  "probabilities": [0.17, 0.83],
  "confidence": 0.64,
  "evidence_summary": "Brief summary of the strongest respondent-level evidence.",
  "notes": "Brief explanation of the respondent-level probability estimates."
}

The probabilities array must contain one number for each held-out option, in the same order as the options. Each probability must be between 0 and 1. The probabilities should sum to 1. Confidence is your uncertainty-calibrated confidence, from 0 to 1, in the individual-level distribution you provided."""
    return base + """
Return only valid JSON. Do not include markdown fences, prose, or comments.

The JSON must have exactly this shape:
{
  "probabilities": [0.17, 0.83],
  "notes": "Brief explanation of the respondent-level probability estimates."
}

The probabilities array must contain one number for each held-out option, in the same order as the options. Each probability must be between 0 and 1. The probabilities should sum to 1."""


def chunked_job_id(prefix: str, chunk_index: int) -> str:
    clean_prefix = slug_id(prefix).lower() or "twin_holdout"
    return f"{clean_prefix}_chunk_{chunk_index:03d}"


def result_chunk_label(path: Path, index: int) -> str:
    match = re.search(r"chunk[_-](\d+)", path.name)
    if match:
        return f"chunk_{int(match.group(1)):03d}"
    return f"chunk_{index:03d}"


def build_edsl_digital_twin_job_dict(survey_name: str, args: Any, deps: DigitalTwinJobBuilderDeps) -> dict[str, Any]:
    sdir = deps.require_survey(survey_name)
    questions = read_jsonl(sdir / "questions.jsonl")
    extra_specs = extra_heldout_question_specs(args)
    if extra_specs:
        existing_names = {question["question_name"] for question in questions}
        questions.extend([spec for spec in extra_specs if spec["question_name"] not in existing_names])
    question_by_name = {question["question_name"]: question for question in questions}
    if args.balance_actual and args.stratify_actual:
        raise ZwillError("invalid_input", "Use only one of --balance-actual or --stratify-actual.")
    heldout_names = selected_heldout_question_names(args, questions)
    unsupported = [
        name
        for name in heldout_names
        if question_by_name[name].get("question_type") != "multiple_choice" or not question_by_name[name].get("question_options")
    ]
    if unsupported:
        raise ZwillError(
            "invalid_input",
            "Digital twin probability export only supports multiple-choice held-out questions with expanded options.",
            context={"unsupported_questions": unsupported},
            hint="Import codebooks first so question_options contain human-readable labels.",
        )

    context_args = type("ContextArgs", (), {})()
    context_args.question = args.context_question
    context_args.questions = args.context_questions
    context_args.exclude_question = args.exclude_context_question or []
    context_question_names = deps.selected_question_names(context_args, questions)
    raw_leakage_exclusions = target_specific_leakage_exclusions(args)
    unknown_exclusion_targets = sorted(set(raw_leakage_exclusions) - set(heldout_names))
    if unknown_exclusion_targets:
        raise ZwillError(
            "invalid_input",
            "Leakage exclusions reference held-out targets that are not selected.",
            context={"unknown_targets": unknown_exclusion_targets, "heldout_questions": heldout_names},
        )
    known_question_names = {question["question_name"] for question in questions}
    # Map each rank_task_id -> its item question names, so a whole battery can be
    # excluded with one `target:<rank_task_id>` flag.
    rank_task_items: dict[str, list[str]] = defaultdict(list)
    for question in questions:
        task_id = str(question.get("rank_task_id") or "")
        if task_id:
            rank_task_items[task_id].append(str(question["question_name"]))
    leakage_exclusions = resolve_leakage_exclusion_patterns(
        raw_leakage_exclusions, known_question_names, dict(rank_task_items)
    )
    all_respondent_ids = [row["respondent_id"] for row in read_jsonl(sdir / "respondents.jsonl")]
    if not all_respondent_ids:
        all_respondent_ids = sorted({row["respondent_id"] for row in read_jsonl(sdir / "answers.jsonl")})
    respondent_ids = deps.respondent_selection(args, all_respondent_ids)

    answer_by_respondent: dict[str, dict[str, str]] = defaultdict(dict)
    for answer in read_jsonl(sdir / "answers.jsonl"):
        if answer.get("answer") is None:
            continue
        answer_by_respondent[answer["respondent_id"]][answer["question"]] = answer["answer"]
    counts_by_question = answer_commonness_by_question(answer_by_respondent)
    if args.complete_cases:
        required_questions = set(heldout_names) | set(context_question_names)
        respondent_ids = [
            respondent_id
            for respondent_id in respondent_ids
            if required_questions.issubset(answer_by_respondent.get(respondent_id, {}))
        ]

    context_file = deps.context_path(sdir)
    context_text = context_file.read_text().strip() if context_file.exists() else ""
    Jobs, Model, ModelList, QuestionFreeText, Scenario, ScenarioList, Survey = deps.load_edsl_job_classes()
    prompt_variant = getattr(args, "prompt_variant", "raw") or "raw"
    question_text = digital_twin_question_text(prompt_variant)

    twin_question = QuestionFreeText(
        question_name=args.job_question_name,
        question_text=question_text,
    )
    all_twin_material = deps.load_twin_material(args)
    scenarios = []
    skipped_missing_heldout = []
    for heldout_name in heldout_names:
        heldout = question_by_name[heldout_name]
        heldout_respondent_ids = respondent_ids
        if args.balance_actual:
            heldout_respondent_ids = balanced_by_actual(
                respondent_ids,
                answer_by_respondent,
                heldout_name,
                args.sample_respondents,
                args.seed,
            )
        elif args.stratify_actual:
            heldout_respondent_ids = stratified_by_actual(
                respondent_ids,
                answer_by_respondent,
                heldout_name,
                args.sample_respondents,
                args.seed,
            )
        for respondent_id in heldout_respondent_ids:
            respondent_answers = answer_by_respondent.get(respondent_id, {})
            actual_answer = respondent_answers.get(heldout_name)
            if actual_answer is None:
                if not getattr(args, "allow_missing_actual", False):
                    skipped_missing_heldout.append({"respondent_id": respondent_id, "heldout_question": heldout_name})
                    continue
            target_exclusions = leakage_exclusions.get(heldout_name, set())
            target_context_question_names = [
                question_name
                for question_name in context_question_names
                if question_name not in target_exclusions
            ]
            selected_context = select_context_questions(
                respondent_answers,
                target_context_question_names,
                heldout_name,
                args.context_question_count,
            )
            observed_answers = [
                {
                    "question_name": question_name,
                    "question_text": expand_question_text_fields(
                        question_by_name[question_name]["question_text"],
                        respondent_answers,
                        question_by_name,
                    ),
                    "question_options": deps.context_question_options(question_by_name[question_name]),
                    "answer": respondent_answers[question_name],
                }
                for question_name in selected_context
            ]
            agent_material = (
                deps.select_agent_material(sdir, [respondent_id], args)
                if getattr(args, "include_agent_material", False)
                else []
            )
            agent_material_text = deps.format_agent_material(
                agent_material,
                getattr(args, "max_agent_material_chars", None),
            )
            twin_material = deps.matching_twin_material(
                all_twin_material,
                survey_name=survey_name,
                heldout_question=heldout_name,
                respondent_id=respondent_id,
            )
            twin_material_text = deps.format_twin_material(
                twin_material,
                getattr(args, "max_twin_material_chars", None),
            )
            observed_lines = []
            for observed in observed_answers:
                observed_lines.append(
                    "\n".join(
                        [
                            f"Question: {observed['question_name']}",
                            f"Text: {observed['question_text']}",
                            "Options: " + "; ".join(observed["question_options"]),
                            f"Respondent answered: {observed['answer']}",
                            *(
                                [answer_commonness_text(observed["question_name"], observed["answer"], counts_by_question)]
                                if prompt_variant == "answer-commonness-confidence"
                                else []
                            ),
                        ]
                    )
                )
            option_keys = [deps.option_key(index) for index, _ in enumerate(heldout["question_options"])]
            option_lines = [f"{key}: {option}" for key, option in zip(option_keys, heldout["question_options"])]
            heldout_question_text = expand_question_text_fields(
                heldout["question_text"],
                respondent_answers,
                question_by_name,
            )
            scenarios.append(
                Scenario(
                    {
                        "survey_name": survey_name,
                        "survey_context": context_text,
                        "respondent_id": respondent_id,
                        "heldout_question_name": heldout["question_name"],
                        "heldout_question_text": heldout_question_text,
                        "heldout_options": heldout["question_options"],
                        "heldout_option_keys": option_keys,
                        "heldout_options_text": "\n".join(option_lines),
                        "actual_answer": actual_answer,
                        "agent_material": agent_material,
                        "agent_material_text": agent_material_text,
                        "twin_material": twin_material,
                        "twin_material_text": twin_material_text,
                        "observed_answers": observed_answers,
                        "observed_answers_text": "\n\n".join(observed_lines) if observed_lines else "No observed answers provided.",
                        "leakage_exclusions": sorted(target_exclusions),
                    }
                )
            )

    if not scenarios:
        raise ZwillError(
            "invalid_input",
            "No digital twin scenarios could be built.",
            context={"skipped_missing_heldout": skipped_missing_heldout[:10], "skipped_count": len(skipped_missing_heldout)},
        )

    model_params = deps.parse_model_params(args)
    job = Jobs(
        survey=Survey(questions=[twin_question]),
        scenarios=ScenarioList(scenarios),
        models=ModelList(
            [
                Model(
                    model_name=model_name,
                    service_name=service_name,
                    **deps.model_kwargs_for(model_name, service_name, model_params),
                )
                for model_name, service_name in deps.parse_model_specs(args)
            ]
        ),
    )
    data = job.to_dict()
    data["zwill"] = {
        "digital_twin_job_id": digital_twin_job_id_from_job(data),
        "heldout_questions": heldout_names,
        "context_question_count": args.context_question_count,
        "sample_respondents": args.sample_respondents,
        "seed": args.seed,
        "complete_cases": args.complete_cases,
        "balance_actual": args.balance_actual,
        "stratify_actual": args.stratify_actual,
        "include_agent_material": getattr(args, "include_agent_material", False),
        "agent_material_kinds": sorted(deps.selected_agent_material_kinds(args)),
        "agent_material_tags": sorted(deps.selected_agent_material_tags(args)),
        "max_agent_material_chars": getattr(args, "max_agent_material_chars", None),
        "twin_material_paths": deps.twin_material_paths(args),
        "max_twin_material_chars": getattr(args, "max_twin_material_chars", None),
        "twin_material_count": len(all_twin_material),
        "extra_heldout_question_count": len(extra_specs),
        "allow_missing_actual": getattr(args, "allow_missing_actual", False),
        "leakage_exclusions": {target: sorted(values) for target, values in sorted(leakage_exclusions.items())},
        "prompt_variant": prompt_variant,
        "scenario_count": len(scenarios),
        "skipped_missing_heldout_count": len(skipped_missing_heldout),
    }
    return data
