from __future__ import annotations

import csv
import html
import json
import math
import os
import time
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


def cosine(a: list[float], b: list[float]) -> float:
    dot = sum(x * y for x, y in zip(a, b))
    na = math.sqrt(sum(x * x for x in a))
    nb = math.sqrt(sum(y * y for y in b))
    return dot / (na * nb) if na and nb else 0.0


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    outdir = root / "survey_results_held_out_questions_analysis"
    outdir.mkdir(exist_ok=True)
    workbook = root / "survey_results_held_out_questions.xlsx"
    key = os.environ.get("OPENAI_API_KEY")
    if not key:
        raise SystemExit("OPENAI_API_KEY missing")

    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["Questions"]
    header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: list[dict[str, object]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(header, raw))
        qid = str(rec.get("Question code") or "").strip()
        qtext = str(rec.get("Question text") or "").strip()
        if not qid or not qtext:
            continue
        labels = str(rec.get("Answer value labels") or "").strip()
        options: list[str] = []
        for i in range(1, 21):
            val = rec.get(f"Answer option {i}")
            if val is not None and str(val).strip():
                options.append(str(val).strip())
        embed_text = qtext
        if labels:
            embed_text += f"\nAnswer value labels: {labels}"
        if options:
            embed_text += "\nAnswer options: " + " | ".join(options)
        rows.append(
            {
                "question": qid,
                "text": qtext,
                "labels": labels,
                "options": options,
                "embed_text": embed_text,
            }
        )

    model = "text-embedding-3-large"
    payload = json.dumps(
        {"model": model, "input": [str(r["embed_text"]) for r in rows]}
    ).encode("utf-8")
    req = urllib.request.Request(
        "https://api.openai.com/v1/embeddings",
        data=payload,
        headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
        method="POST",
    )

    embeddings: list[list[float]] = []
    last_err = None
    for attempt in range(4):
        try:
            with urllib.request.urlopen(req, timeout=90) as resp:
                data = json.loads(resp.read().decode("utf-8"))
            embeddings = [
                item["embedding"] for item in sorted(data["data"], key=lambda x: x["index"])
            ]
            break
        except urllib.error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            last_err = f"HTTP {exc.code}: {body[:500]}"
            if exc.code not in {429, 500, 502, 503, 504}:
                raise SystemExit(last_err)
        except Exception as exc:  # noqa: BLE001 - preserve API failure detail.
            last_err = repr(exc)
        time.sleep(2**attempt)
    if not embeddings:
        raise SystemExit(f"Embedding request failed: {last_err}")

    n = len(rows)
    sim = [[cosine(embeddings[i], embeddings[j]) for j in range(n)] for i in range(n)]

    matrix_path = outdir / "question_embedding_cosine_similarity_openai.csv"
    with matrix_path.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["question"] + [str(r["question"]) for r in rows])
        for i, row in enumerate(rows):
            writer.writerow([row["question"]] + [f"{sim[i][j]:.6f}" for j in range(n)])

    pairs = []
    for i in range(n):
        for j in range(i + 1, n):
            pairs.append(
                {
                    "question_i": str(rows[i]["question"]),
                    "question_j": str(rows[j]["question"]),
                    "cosine_similarity": sim[i][j],
                    "text_i": str(rows[i]["text"]),
                    "text_j": str(rows[j]["text"]),
                }
            )
    pairs.sort(key=lambda r: r["cosine_similarity"], reverse=True)

    pairs_path = outdir / "question_embedding_cosine_similarity_pairs_openai.csv"
    with pairs_path.open("w", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "question_i",
                "question_j",
                "cosine_similarity",
                "text_i",
                "text_j",
            ],
        )
        writer.writeheader()
        for row in pairs:
            writer.writerow({**row, "cosine_similarity": f"{row['cosine_similarity']:.6f}"})

    meta_path = outdir / "question_embedding_metadata_openai.json"
    meta_path.write_text(
        json.dumps(
            {
                "model": model,
                "question_count": n,
                "source": str(workbook),
                "embedding_text": "question text + answer value labels + answer options",
                "matrix_csv": str(matrix_path),
                "pairs_csv": str(pairs_path),
                "top_pairs": [
                    {
                        "question_i": row["question_i"],
                        "question_j": row["question_j"],
                        "cosine_similarity": round(row["cosine_similarity"], 6),
                    }
                    for row in pairs[:20]
                ],
            },
            indent=2,
            ensure_ascii=False,
        )
    )

    labels = [str(r["question"]) for r in rows]
    min_val = min(sim[i][j] for i in range(n) for j in range(i + 1, n))
    max_val = max(sim[i][j] for i in range(n) for j in range(i + 1, n))

    def color(value: float) -> str:
        t = 0 if max_val == min_val else (value - min_val) / (max_val - min_val)
        if t < 0.55:
            u = t / 0.55
            c1 = (247, 251, 255)
            c2 = (43, 108, 176)
        else:
            u = (t - 0.55) / 0.45
            c1 = (43, 108, 176)
            c2 = (185, 28, 28)
        rgb = tuple(round(c1[k] + (c2[k] - c1[k]) * u) for k in range(3))
        return "#%02x%02x%02x" % rgb

    cell = 16
    left = 150
    top = 150
    svg_w = left + n * cell + 30
    svg_h = top + n * cell + 90
    parts = [
        f"""<!doctype html>
<html><head><meta charset="utf-8"><title>Question Embedding Similarity Heatmap</title>
<style>
body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;margin:28px;color:#17202a;background:#fff;}}
h1{{margin:0 0 6px;font-size:28px;}}
.summary{{color:#52606d;max-width:1050px;margin-bottom:18px;line-height:1.45;}}
.wrap{{overflow:auto;border:1px solid #d8dee6;border-radius:8px;padding:12px;}}
.cell{{stroke:#fff;stroke-width:.7;}}
.axis{{font-size:10px;fill:#334e68;}}
.tick:hover{{font-weight:700;}}
.cell:hover{{stroke:#111827;stroke-width:1.4;}}
.table{{border-collapse:collapse;margin-top:22px;font-size:13px;max-width:1200px;}}
.table th,.table td{{border:1px solid #d8dee6;padding:6px 8px;vertical-align:top;}}
.table th{{background:#f3f5f7;}}
.code{{font-family:ui-monospace,SFMono-Regular,Menlo,monospace;}}
</style></head><body>
<h1>Question Embedding Cosine Similarity</h1>
<div class="summary">OpenAI <span class="code">{html.escape(model)}</span> embeddings of question text plus answer labels/options. The heatmap shows the upper triangle only; darker red means more similar. Questions included: {n}. Similarity range: {min_val:.3f} to {max_val:.3f}.</div>
<div class="wrap"><svg width="{svg_w}" height="{svg_h}" role="img" aria-label="Upper triangle heatmap of question embedding cosine similarity">
"""
    ]
    for idx, label in enumerate(labels):
        x = left + idx * cell + cell / 2
        y = top - 6
        parts.append(
            f'<text class="axis tick" transform="translate({x:.1f},{y:.1f}) rotate(-60)" text-anchor="start">{html.escape(label)}</text>\n'
        )
        parts.append(
            f'<text class="axis tick" x="{left - 8}" y="{top + idx * cell + cell * 0.72:.1f}" text-anchor="end">{html.escape(label)}</text>\n'
        )
    for i in range(n):
        for j in range(i + 1, n):
            value = sim[i][j]
            x = left + j * cell
            y = top + i * cell
            title = (
                f"{rows[i]['question']} x {rows[j]['question']}: {value:.3f}\n"
                f"{rows[i]['text']}\n---\n{rows[j]['text']}"
            )
            parts.append(
                f'<rect class="cell" x="{x}" y="{y}" width="{cell}" height="{cell}" fill="{color(value)}"><title>{html.escape(title)}</title></rect>\n'
            )
    legend_x = left
    legend_y = top + n * cell + 34
    parts.append(f'<text class="axis" x="{legend_x}" y="{legend_y - 10}">low</text>')
    for k in range(120):
        value = min_val + (max_val - min_val) * k / 119
        parts.append(
            f'<rect x="{legend_x + 30 + k}" y="{legend_y - 20}" width="1" height="12" fill="{color(value)}"/>'
        )
    parts.append(f'<text class="axis" x="{legend_x + 158}" y="{legend_y - 10}">high</text>')
    parts.append("</svg></div>")
    parts.append(
        '<h2>Top Similar Question Pairs</h2><table class="table"><thead><tr><th>Rank</th><th>Pair</th><th>Cosine</th><th>Question texts</th></tr></thead><tbody>'
    )
    for rank, row in enumerate(pairs[:25], 1):
        parts.append("<tr>")
        parts.append(
            f'<td>{rank}</td><td class="code">{html.escape(row["question_i"])} x {html.escape(row["question_j"])}</td><td>{row["cosine_similarity"]:.3f}</td>'
        )
        parts.append(
            f'<td><strong>{html.escape(row["question_i"])}</strong>: {html.escape(row["text_i"])}<br><strong>{html.escape(row["question_j"])}</strong>: {html.escape(row["text_j"])}</td>'
        )
        parts.append("</tr>")
    parts.append("</tbody></table></body></html>")
    html_path = outdir / "question_embedding_similarity_heatmap_openai.html"
    html_path.write_text("".join(parts))

    print(
        json.dumps(
            {
                "question_count": n,
                "model": model,
                "html": str(html_path),
                "matrix_csv": str(matrix_path),
                "pairs_csv": str(pairs_path),
                "metadata": str(meta_path),
                "top_pairs": [
                    {
                        "question_i": row["question_i"],
                        "question_j": row["question_j"],
                        "cosine_similarity": round(row["cosine_similarity"], 4),
                    }
                    for row in pairs[:10]
                ],
            },
            indent=2,
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
