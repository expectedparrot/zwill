#!/usr/bin/env python3
"""Generate and export answer-commonness twin predictions for true held-outs."""

from __future__ import annotations

import argparse
import csv
import gzip
import json
import math
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl

from zwill.probability import normalized_probabilities, parse_probability_json


ROOT = Path(__file__).resolve().parents[1]
SURVEY = "survey_results_held_out_questions"
SDIR = ROOT / ".zwill/projects/excel_survey_analysis/surveys" / SURVEY
WORKBOOK = ROOT / "survey_results_held_out_questions.xlsx"
RUN_DIR = ROOT / "excel_true_heldout_answer_commonness"
TEMPLATE_JOB = ROOT / "excel_random_sample_5q_variants/answer_commonness_confidence_chunk_01_job.edsl.json"
TRUE_HELDOUTS = [
    "q13",
    "q39",
    "q49",
    "q50",
    "q51",
    "q52",
    "q53",
    "q54",
    "q55",
    "q56",
    "q57",
    "q58",
    "q59",
    "q71",
]


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    return [json.loads(line) for line in path.read_text().splitlines() if line.strip()]


def read_result(path: Path) -> dict[str, Any]:
    if path.suffix == ".gz":
        with gzip.open(path, "rt") as f:
            return json.load(f)
    return json.loads(path.read_text())


def safe_option_column(option: str, index: int) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "_", option).strip("_").lower()
    slug = slug[:50] or f"option_{index}"
    return f"prob_{index:02d}_{slug}"


def question_options(question: dict[str, Any]) -> list[str]:
    if question.get("question_options"):
        return [str(value) for value in question["question_options"]]
    known = question.get("source", {}).get("known_options")
    if known:
        return [str(value) for value in known]
    return []


def workbook_true_heldouts() -> dict[str, dict[str, Any]]:
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb["Questions"]
    headers = [str(ws.cell(1, column).value) for column in range(1, ws.max_column + 1)]
    by_name = {}
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = dict(zip(headers, row))
        name = str(record.get("Question code") or "").strip()
        if name not in TRUE_HELDOUTS:
            continue
        labels = str(record.get("Answer value labels") or "")
        options = []
        for key in headers:
            if not key.startswith("Answer option "):
                continue
            value = record.get(key)
            if value is None or str(value).strip() == "":
                continue
            raw = str(value).strip()
            if labels == "1=lowest agreement/likelihood/appeal/excitement; 7=highest":
                if raw == "1":
                    option = "1 - lowest agreement/likelihood/appeal/excitement"
                elif raw == "7":
                    option = "7 - highest agreement/likelihood/appeal/excitement"
                else:
                    option = f"{raw} - Likert scale point {raw}"
            else:
                option = raw
            options.append(option)
        by_name[name] = {
            "question_name": name,
            "question_text": str(record["Question text"]),
            "question_options": options,
            "source": {
                "raw_id": "source_workbook",
                "note": f"Questions sheet row {row_number}. Source workbook held-out question.",
                "answer_value_labels": labels or None,
            },
        }
    missing = [name for name in TRUE_HELDOUTS if name not in by_name]
    if missing:
        raise SystemExit(f"Missing held-out question definitions in workbook: {missing}")
    return by_name


def commonness_label(probability: float) -> str:
    if probability < 0.05:
        return "very rare"
    if probability < 0.15:
        return "rare"
    if probability < 0.35:
        return "moderately common"
    return "common"


def load_context() -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]], dict[str, dict[str, str]], dict[str, Counter]]:
    respondents = read_jsonl(SDIR / "respondents.jsonl")
    questions = {row["question_name"]: row for row in read_jsonl(SDIR / "questions.jsonl")}
    answers: dict[str, dict[str, str]] = defaultdict(dict)
    counts: dict[str, Counter] = defaultdict(Counter)
    for row in read_jsonl(SDIR / "answers.jsonl"):
        respondent = row["respondent_id"]
        question = row["question"]
        answer = str(row["answer"])
        answers[respondent][question] = answer
        counts[question][answer] += 1
    return respondents, questions, answers, counts


def format_options(options: list[str]) -> str:
    return "\n".join(f"{chr(97 + index)}: {option}" for index, option in enumerate(options))


def observed_text(
    respondent_answers: dict[str, str],
    questions: dict[str, dict[str, Any]],
    counts: dict[str, Counter],
) -> tuple[list[dict[str, Any]], str]:
    observed = []
    blocks = []
    for question_name in questions:
        if question_name not in respondent_answers:
            continue
        question = questions[question_name]
        options = question_options(question)
        answer = respondent_answers[question_name]
        observed.append(
            {
                "question_name": question_name,
                "question_text": question["question_text"],
                "question_options": options,
                "answer": answer,
            }
        )
        lines = [
            f"Question: {question_name}",
            f"Text: {question['question_text']}",
        ]
        if options:
            lines.append(f"Options: {'; '.join(options)}")
        lines.append(f"Respondent answered: {answer}")
        if answer in counts[question_name]:
            numerator = counts[question_name][answer]
            denominator = sum(counts[question_name].values())
            probability = numerator / denominator if denominator else 0.0
            lines.append(
                "Answer commonness for this context question: "
                f"{probability:.3f} ({numerator}/{denominator}), {commonness_label(probability)}."
            )
        else:
            lines.append("Answer commonness: not available as a single-choice categorical statistic.")
        blocks.append("\n".join(lines))
    return observed, "\n\n".join(blocks)


def build_jobs(chunk_size: int) -> None:
    RUN_DIR.mkdir(exist_ok=True)
    template = json.loads(TEMPLATE_JOB.read_text())
    template["zwill"] = {
        "survey": SURVEY,
        "digital_twin_job_id": "true_heldout_answer_commonness_v1",
        "variant": "answer_commonness_confidence",
        "question_set": TRUE_HELDOUTS,
    }
    respondents, context_questions, answers, counts = load_context()
    heldouts = workbook_true_heldouts()
    survey_context = (SDIR / "context.md").read_text()

    scenarios = []
    for question_name in TRUE_HELDOUTS:
        heldout = heldouts[question_name]
        options = heldout["question_options"]
        for respondent in respondents:
            respondent_id = respondent["respondent_id"]
            observed, observed_answers_text = observed_text(answers[respondent_id], context_questions, counts)
            scenarios.append(
                {
                    "survey_name": SURVEY,
                    "survey_context": survey_context,
                    "respondent_id": respondent_id,
                    "respondent_source_row": respondent.get("metadata", {}).get("source_row"),
                    "heldout_question_name": question_name,
                    "heldout_question_text": heldout["question_text"],
                    "heldout_options": options,
                    "heldout_option_keys": [chr(97 + index) for index in range(len(options))],
                    "heldout_options_text": format_options(options),
                    "actual_answer": None,
                    "agent_material": [],
                    "agent_material_text": "No non-survey agent material provided.",
                    "twin_material": [],
                    "twin_material_text": "No supplemental twin material supplied.",
                    "observed_answers": observed,
                    "observed_answers_text": observed_answers_text,
                    "edsl_version": "1.0.8.dev1",
                    "edsl_class_name": "Scenario",
                    "true_heldout_variant": "answer_commonness_confidence",
                }
            )

    manifest = {
        "survey": SURVEY,
        "variant": "answer_commonness_confidence",
        "questions": TRUE_HELDOUTS,
        "respondent_count": len(respondents),
        "scenario_count": len(scenarios),
        "chunk_size": chunk_size,
        "workbook": str(WORKBOOK.relative_to(ROOT)),
        "heldout_questions": heldouts,
    }
    (RUN_DIR / "manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + "\n")

    for old in RUN_DIR.glob("chunk_*_job.edsl.json"):
        old.unlink()
    for chunk_index, start in enumerate(range(0, len(scenarios), chunk_size), start=1):
        chunk = scenarios[start : start + chunk_size]
        job = dict(template)
        job["scenarios"] = chunk
        job["zwill"] = dict(template["zwill"])
        job["zwill"]["digital_twin_job_id"] = f"true_heldout_answer_commonness_chunk_{chunk_index:03d}_v1"
        path = RUN_DIR / f"chunk_{chunk_index:03d}_job.edsl.json"
        path.write_text(json.dumps(job, ensure_ascii=False, separators=(",", ":")) + "\n")
    print(f"Wrote {len(list(RUN_DIR.glob('chunk_*_job.edsl.json')))} job chunks to {RUN_DIR}")


def extract_answer_payload(row: dict[str, Any]) -> tuple[dict[str, Any] | None, str | None]:
    answer = row.get("answer", {})
    raw = None
    if isinstance(answer, dict):
        raw = answer.get("response_probabilities")
        if raw is None and answer:
            raw = next(iter(answer.values()))
    return parse_probability_json(raw)


def export_csv() -> None:
    manifest = json.loads((RUN_DIR / "manifest.json").read_text())
    result_paths = sorted(RUN_DIR.glob("chunk_*_results.json.gz"))
    long_path = RUN_DIR / "true_heldout_answer_commonness_predictions_long.csv"
    wide_path = RUN_DIR / "true_heldout_answer_commonness_predictions_wide.csv"
    issue_path = RUN_DIR / "true_heldout_answer_commonness_issues.jsonl"
    long_rows = []
    wide_rows = []
    issues = []
    seen = set()
    for path in result_paths:
        results = read_result(path)
        job_id = results.get("zwill", {}).get("digital_twin_job_id")
        for index, row in enumerate(results.get("data", [])):
            scenario = row.get("scenario", {})
            options = [str(option) for option in scenario.get("heldout_options", [])]
            parsed, parse_error = extract_answer_payload(row)
            probabilities = parsed.get("probabilities") if isinstance(parsed, dict) else None
            notes = parsed.get("notes") if isinstance(parsed, dict) else None
            confidence = parsed.get("confidence") if isinstance(parsed, dict) else None
            evidence_summary = parsed.get("evidence_summary") if isinstance(parsed, dict) else None
            normalized = None
            probability_sum = None
            error = parse_error
            if isinstance(probabilities, list):
                try:
                    values = [float(value) for value in probabilities]
                    normalized, probability_sum, norm_error = normalized_probabilities(values, len(options))
                    error = error or norm_error
                except (TypeError, ValueError):
                    error = error or "invalid_probability_value"
            else:
                error = error or "missing_probabilities"
            key = (scenario.get("respondent_id"), scenario.get("heldout_question_name"))
            if key in seen:
                error = error or "duplicate_prediction"
            if error:
                issues.append(
                    {
                        "results_path": str(path.relative_to(ROOT)),
                        "job_id": job_id,
                        "row": index,
                        "respondent_id": scenario.get("respondent_id"),
                        "heldout_question": scenario.get("heldout_question_name"),
                        "error": error,
                    }
                )
                continue
            seen.add(key)
            probabilities_by_option = {option: normalized[position] for position, option in enumerate(options)}
            top_choice = max(probabilities_by_option.items(), key=lambda item: item[1])[0]
            base = {
                "respondent_id": scenario.get("respondent_id"),
                "respondent_source_row": scenario.get("respondent_source_row"),
                "heldout_question": scenario.get("heldout_question_name"),
                "heldout_question_text": scenario.get("heldout_question_text"),
                "job_id": job_id,
                "model": row.get("model", {}).get("model"),
                "service": row.get("model", {}).get("inference_service"),
                "raw_probability_sum": probability_sum,
                "top_choice": top_choice,
                "top_probability": probabilities_by_option[top_choice],
                "confidence": confidence,
                "evidence_summary": evidence_summary,
                "notes": notes,
            }
            wide = dict(base)
            for position, option in enumerate(options, start=1):
                probability = probabilities_by_option[option]
                long_rows.append(
                    {
                        **base,
                        "option_index": position,
                        "option_label": option,
                        "probability": probability,
                    }
                )
                wide[f"option_{position:02d}_label"] = option
                wide[safe_option_column(option, position)] = probability
            wide_rows.append(wide)

    expected = manifest["scenario_count"]
    missing = expected - len(seen)
    if missing:
        issues.append({"error": "missing_predictions", "missing_count": missing, "expected": expected, "seen": len(seen)})

    if long_rows:
        with long_path.open("w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=list(long_rows[0]))
            writer.writeheader()
            writer.writerows(long_rows)
    if wide_rows:
        all_fields = []
        for row in wide_rows:
            for field in row:
                if field not in all_fields:
                    all_fields.append(field)
        with wide_path.open("w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=all_fields)
            writer.writeheader()
            writer.writerows(wide_rows)
    with issue_path.open("w") as f:
        for issue in issues:
            f.write(json.dumps(issue, ensure_ascii=False) + "\n")

    summary = {
        "result_files": len(result_paths),
        "expected_predictions": expected,
        "exported_predictions": len(seen),
        "long_rows": len(long_rows),
        "wide_rows": len(wide_rows),
        "issue_count": len(issues),
        "long_csv": str(long_path.relative_to(ROOT)),
        "wide_csv": str(wide_path.relative_to(ROOT)),
        "issues": str(issue_path.relative_to(ROOT)),
    }
    (RUN_DIR / "export_summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False) + "\n")
    print(json.dumps(summary, indent=2, ensure_ascii=False))


def main() -> None:
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="command", required=True)
    build = sub.add_parser("build-jobs")
    build.add_argument("--chunk-size", type=int, default=75)
    sub.add_parser("export-csv")
    args = parser.parse_args()
    if args.command == "build-jobs":
        build_jobs(args.chunk_size)
    elif args.command == "export-csv":
        export_csv()


if __name__ == "__main__":
    main()
